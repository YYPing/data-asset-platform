"""
证书解析工具 - 支持多种格式的证书解析
"""
import io
import re
from typing import Dict, Any, Optional, List
from datetime import date, datetime
import json

try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

from app.utils.ocr_processor import get_ocr_processor


class CertificateParser:
    """证书解析器 - 支持PDF、图片、Excel格式"""
    
    def __init__(self):
        """初始化证书解析器"""
        self.ocr_processor = get_ocr_processor()
    
    def parse_from_file(self, file_content: bytes, file_format: str, 
                       filename: str = "") -> Dict[str, Any]:
        """
        从文件解析证书信息
        
        Args:
            file_content: 文件内容
            file_format: 文件格式（pdf/jpg/jpeg/png/excel/xlsx）
            filename: 文件名（可选）
            
        Returns:
            dict: 解析出的证书信息
        """
        file_format = file_format.lower().replace('.', '')
        
        if file_format == 'pdf':
            return self.parse_pdf(file_content)
        elif file_format in ['jpg', 'jpeg', 'png']:
            return self.parse_image(file_content)
        elif file_format in ['excel', 'xlsx', 'xls']:
            return self.parse_excel(file_content)
        else:
            raise ValueError(f"不支持的文件格式: {file_format}")
    
    def parse_pdf(self, file_content: bytes) -> Dict[str, Any]:
        """
        解析PDF证书
        
        Args:
            file_content: PDF文件内容
            
        Returns:
            dict: 解析出的证书信息
        """
        # 提取文本
        text = self.ocr_processor.extract_text_from_pdf(file_content)
        
        # 解析证书信息
        info = self.ocr_processor.parse_certificate_info(text)
        
        # 添加原始文本
        info['raw_text'] = text
        info['parse_method'] = 'pdf_text_extraction'
        
        return info
    
    def parse_image(self, file_content: bytes) -> Dict[str, Any]:
        """
        解析图片证书（OCR）
        
        Args:
            file_content: 图片文件内容
            
        Returns:
            dict: 解析出的证书信息
        """
        # OCR识别
        text = self.ocr_processor.extract_text_from_image(file_content)
        
        # 解析证书信息
        info = self.ocr_processor.parse_certificate_info(text)
        
        # 添加原始文本
        info['raw_text'] = text
        info['parse_method'] = 'ocr'
        
        return info
    
    def parse_excel(self, file_content: bytes) -> List[Dict[str, Any]]:
        """
        解析Excel证书（批量导入）
        
        Args:
            file_content: Excel文件内容
            
        Returns:
            list: 证书信息列表
        """
        if not OPENPYXL_AVAILABLE and not PANDAS_AVAILABLE:
            raise RuntimeError("需要安装openpyxl或pandas库来解析Excel文件")
        
        certificates = []
        
        try:
            # 优先使用pandas（更快）
            if PANDAS_AVAILABLE:
                df = pd.read_excel(io.BytesIO(file_content))
                
                # 标准化列名
                df.columns = [self._normalize_column_name(col) for col in df.columns]
                
                # 转换为字典列表
                for _, row in df.iterrows():
                    cert_info = self._parse_excel_row(row.to_dict())
                    if cert_info:
                        certificates.append(cert_info)
            
            # 备用方案：使用openpyxl
            elif OPENPYXL_AVAILABLE:
                wb = load_workbook(io.BytesIO(file_content))
                ws = wb.active
                
                # 读取表头
                headers = []
                for cell in ws[1]:
                    headers.append(self._normalize_column_name(cell.value or ""))
                
                # 读取数据行
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    cert_info = self._parse_excel_row(row_dict)
                    if cert_info:
                        certificates.append(cert_info)
        
        except Exception as e:
            raise RuntimeError(f"解析Excel文件失败: {str(e)}")
        
        return certificates
    
    def _normalize_column_name(self, col_name: str) -> str:
        """
        标准化列名
        
        Args:
            col_name: 原始列名
            
        Returns:
            str: 标准化后的列名
        """
        col_name = str(col_name).strip().lower()
        
        # 列名映射
        column_mapping = {
            '证书编号': 'certificate_no',
            '编号': 'certificate_no',
            'certificate_no': 'certificate_no',
            'cert_no': 'certificate_no',
            'no': 'certificate_no',
            
            '证书名称': 'certificate_name',
            '名称': 'certificate_name',
            'certificate_name': 'certificate_name',
            'name': 'certificate_name',
            
            '证书类型': 'certificate_type',
            '类型': 'certificate_type',
            'certificate_type': 'certificate_type',
            'type': 'certificate_type',
            
            '持有人': 'holder_name',
            '持有人姓名': 'holder_name',
            '姓名': 'holder_name',
            'holder_name': 'holder_name',
            'holder': 'holder_name',
            
            '身份证号': 'holder_id_number',
            '证件号': 'holder_id_number',
            '统一社会信用代码': 'holder_id_number',
            'id_number': 'holder_id_number',
            
            '颁发机构': 'issuing_authority',
            '发证机关': 'issuing_authority',
            'issuing_authority': 'issuing_authority',
            'authority': 'issuing_authority',
            
            '颁发日期': 'issue_date',
            '发证日期': 'issue_date',
            'issue_date': 'issue_date',
            
            '有效期': 'expiry_date',
            '有效期至': 'expiry_date',
            '到期日期': 'expiry_date',
            'expiry_date': 'expiry_date',
            'valid_until': 'expiry_date',
            
            '备注': 'notes',
            '说明': 'notes',
            'notes': 'notes',
            'remark': 'notes',
        }
        
        return column_mapping.get(col_name, col_name)
    
    def _parse_excel_row(self, row_dict: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """
        解析Excel行数据
        
        Args:
            row_dict: 行数据字典
            
        Returns:
            dict: 证书信息，如果行无效返回None
        """
        # 必须有证书编号
        if 'certificate_no' not in row_dict or not row_dict['certificate_no']:
            return None
        
        cert_info = {
            'certificate_no': str(row_dict.get('certificate_no', '')).strip(),
            'certificate_name': str(row_dict.get('certificate_name', '')).strip() or None,
            'certificate_type': self._parse_certificate_type(row_dict.get('certificate_type')),
            'holder_name': str(row_dict.get('holder_name', '')).strip() or None,
            'holder_id_number': str(row_dict.get('holder_id_number', '')).strip() or None,
            'issuing_authority': str(row_dict.get('issuing_authority', '')).strip() or None,
            'issue_date': self._parse_date_value(row_dict.get('issue_date')),
            'expiry_date': self._parse_date_value(row_dict.get('expiry_date')),
            'notes': str(row_dict.get('notes', '')).strip() or None,
            'parse_method': 'excel',
        }
        
        return cert_info
    
    def _parse_certificate_type(self, type_value: Any) -> str:
        """
        解析证书类型
        
        Args:
            type_value: 类型值
            
        Returns:
            str: 标准化的证书类型
        """
        if not type_value:
            return 'registration'
        
        type_str = str(type_value).strip().lower()
        
        # 类型映射
        type_mapping = {
            '登记证书': 'registration',
            '数据资产登记证书': 'registration',
            'registration': 'registration',
            
            '合规评估证书': 'compliance',
            '合规证书': 'compliance',
            'compliance': 'compliance',
            
            '价值评估证书': 'value_assessment',
            '评估证书': 'value_assessment',
            'value_assessment': 'value_assessment',
            'valuation': 'value_assessment',
            
            '权属确认证书': 'ownership',
            '权属证书': 'ownership',
            'ownership': 'ownership',
            
            '质量认证证书': 'quality',
            '质量证书': 'quality',
            'quality': 'quality',
        }
        
        return type_mapping.get(type_str, 'registration')
    
    def _parse_date_value(self, date_value: Any) -> Optional[date]:
        """
        解析日期值
        
        Args:
            date_value: 日期值（可能是字符串、datetime或date）
            
        Returns:
            date: 日期对象，解析失败返回None
        """
        if not date_value:
            return None
        
        # 如果已经是date对象
        if isinstance(date_value, date):
            return date_value
        
        # 如果是datetime对象
        if isinstance(date_value, datetime):
            return date_value.date()
        
        # 如果是字符串
        if isinstance(date_value, str):
            date_str = date_value.strip()
            
            # 清理日期字符串
            date_str = date_str.replace('年', '-').replace('月', '-').replace('日', '')
            date_str = date_str.replace('/', '-')
            
            # 尝试多种日期格式
            date_formats = [
                '%Y-%m-%d',
                '%Y%m%d',
                '%Y.%m.%d',
                '%d/%m/%Y',
                '%m/%d/%Y',
            ]
            
            for fmt in date_formats:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    return dt.date()
                except ValueError:
                    continue
        
        return None
    
    def validate_certificate_data(self, cert_data: Dict[str, Any]) -> tuple[bool, List[str]]:
        """
        验证证书数据的完整性
        
        Args:
            cert_data: 证书数据
            
        Returns:
            tuple: (是否有效, 错误消息列表)
        """
        errors = []
        
        # 必填字段检查
        required_fields = ['certificate_no', 'issuing_authority', 'issue_date']
        
        for field in required_fields:
            if not cert_data.get(field):
                field_names = {
                    'certificate_no': '证书编号',
                    'issuing_authority': '颁发机构',
                    'issue_date': '颁发日期',
                }
                errors.append(f"缺少必填字段: {field_names.get(field, field)}")
        
        # 日期逻辑检查
        issue_date = cert_data.get('issue_date')
        expiry_date = cert_data.get('expiry_date')
        
        if issue_date and expiry_date:
            if isinstance(issue_date, str):
                issue_date = self._parse_date_value(issue_date)
            if isinstance(expiry_date, str):
                expiry_date = self._parse_date_value(expiry_date)
            
            if issue_date and expiry_date and expiry_date < issue_date:
                errors.append("有效期不能早于颁发日期")
        
        # 证书编号格式检查
        cert_no = cert_data.get('certificate_no')
        if cert_no and not re.match(r'^[A-Z0-9\-]+$', str(cert_no), re.IGNORECASE):
            errors.append("证书编号格式不正确（应为字母、数字和连字符的组合）")
        
        return len(errors) == 0, errors
    
    def enrich_certificate_data(self, cert_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        丰富证书数据（添加计算字段）
        
        Args:
            cert_data: 证书数据
            
        Returns:
            dict: 丰富后的证书数据
        """
        enriched = cert_data.copy()
        
        # 计算距离过期天数
        expiry_date = cert_data.get('expiry_date')
        if expiry_date:
            if isinstance(expiry_date, str):
                expiry_date = self._parse_date_value(expiry_date)
            
            if expiry_date:
                today = date.today()
                days_until_expiry = (expiry_date - today).days
                enriched['days_until_expiry'] = days_until_expiry
                
                # 自动设置状态
                if days_until_expiry < 0:
                    enriched['auto_status'] = 'expired'
                elif days_until_expiry <= 30:
                    enriched['auto_status'] = 'expiring'
                else:
                    enriched['auto_status'] = 'valid'
        
        return enriched


# 全局解析器实例
_certificate_parser: Optional[CertificateParser] = None


def get_certificate_parser() -> CertificateParser:
    """
    获取证书解析器单例
    
    Returns:
        CertificateParser: 证书解析器实例
    """
    global _certificate_parser
    
    if _certificate_parser is None:
        _certificate_parser = CertificateParser()
    
    return _certificate_parser
